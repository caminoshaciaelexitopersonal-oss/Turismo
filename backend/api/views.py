from rest_framework import generics, views, viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from dj_rest_auth.registration.views import RegisterView
import asyncio
import threading
from django.contrib.contenttypes.models import ContentType
from django.db import models, transaction
from django.utils import timezone
from asgiref.sync import async_to_sync
import json
from datetime import datetime, timedelta
from itertools import groupby
from operator import attrgetter
from django.db.models.functions import TruncDay
from django.db.models import Count
from .models import (
    CustomUser,
    PrestadorServicio,
    ImagenGaleria,
    ImagenArtesano,
    DocumentoLegalizacion,
    Publicacion,
    ConsejoConsultivo,
    AtractivoTuristico,
    RutaTuristica,
    ElementoGuardado,
    CategoriaPrestador,
    Video,
    ContenidoMunicipio,
    AgentTask,
    SiteConfiguration,
    MenuItem,
    HomePageComponent,
    AuditLog,
    PaginaInstitucional,
    ImagenAtractivo,
    Artesano,
    RubroArtesano,
    Resena,
    Sugerencia,
    HechoHistorico,
    ScoringRule,
    Notificacion,
    Formulario,
    Pregunta,
    OpcionRespuesta,
    RespuestaUsuario,
    PlantillaVerificacion,
    ItemVerificacion,
    Verificacion,
    RespuestaItemVerificacion
)
from .serializers import (
    GaleriaItemSerializer,
    PaginaInstitucionalSerializer,
    ScoringRuleSerializer,
    NotificacionSerializer,
    PrestadorServicioSerializer,
    PrestadorServicioUpdateSerializer,
    ArtesanoSerializer,
    ArtesanoUpdateSerializer,
    ImagenGaleriaSerializer,
    ImagenArtesanoSerializer,
    DocumentoLegalizacionSerializer,
    PublicacionListSerializer,
    PublicacionDetailSerializer,
    VideoSerializer,
    ConsejoConsultivoSerializer,
    AtractivoTuristicoListSerializer,
    AtractivoTuristicoDetailSerializer,
    AtractivoTuristicoWriteSerializer,
    RutaTuristicaListSerializer,
    RutaTuristicaDetailSerializer,
    LocationSerializer,
    PrestadorRegisterSerializer,
    TuristaRegisterSerializer,
    ArtesanoRegisterSerializer,
    ElementoGuardadoSerializer,
    ElementoGuardadoCreateSerializer,
    CategoriaPrestadorSerializer,
    PrestadorServicioPublicListSerializer,
    PrestadorServicioPublicDetailSerializer,
    RubroArtesanoSerializer,
    ArtesanoPublicListSerializer,
    ArtesanoPublicDetailSerializer,
    AdminPrestadorListSerializer,
    AdminPrestadorDetailSerializer,
    AdminArtesanoListSerializer,
    AdminArtesanoDetailSerializer,
    UsuarioListSerializer,
    ContenidoMunicipioSerializer,
    AgentCommandSerializer,
    AgentTaskSerializer,
    LLMKeysSerializer,
    SiteConfigurationSerializer,
    MenuItemSerializer,
    AdminUserSerializer,
    HomePageComponentSerializer,
    AuditLogSerializer,
    AdminPublicacionSerializer,
    HechoHistoricoSerializer,
    ResenaSerializer,
    ResenaCreateSerializer,
    SugerenciaSerializer,
    SugerenciaAdminSerializer,
    FeedbackProveedorSerializer,
    FelicitacionPublicaSerializer,
    FormularioListSerializer,
    FormularioDetailSerializer,
    PreguntaSerializer,
    OpcionRespuestaSerializer,
    RespuestaUsuarioSerializer,
    RespuestaUsuarioCreateSerializer,
    # Serializers para Verificación de Cumplimiento
    PlantillaVerificacionListSerializer,
    PlantillaVerificacionDetailSerializer,
    VerificacionListSerializer,
    VerificacionDetailSerializer,
    IniciarVerificacionSerializer,
    GuardarVerificacionSerializer,
    AIConfigSerializer,
# Serializers para Capacitaciones
CapacitacionDetailSerializer,
RegistrarAsistenciaSerializer,
AIConfigSerializer
)
from .permissions import (
    IsTurista,
    IsAdminOrFuncionario,
    IsAdmin,
    IsAdminOrFuncionarioForUserManagement,
    IsPrestador,
    IsAdminOrDirectivo,
    CanManageAtractivos
)
from .filters import AuditLogFilter


class FormularioViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión de Formularios dinámicos.
    - Admins/Directivos: Control total (CRUD).
    - Prestadores/Otros: Solo lectura de formularios públicos.
    """
    queryset = Formulario.objects.all().prefetch_related('preguntas__opciones')

    def get_serializer_class(self):
        if self.action == 'list':
            return FormularioListSerializer
        return FormularioDetailSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdminOrDirectivo]
        return super().get_permissions()

    def get_queryset(self):
        user = self.request.user
        if not user.is_staff and not user.is_superuser:
            return self.queryset.filter(es_publico=True)
        return self.queryset


class PreguntaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar las Preguntas de un Formulario.
    Anidado bajo /formularios/{formulario_pk}/preguntas/
    """
    serializer_class = PreguntaSerializer
    permission_classes = [IsAdminOrDirectivo]

    def get_queryset(self):
        return Pregunta.objects.filter(formulario_id=self.kwargs['formulario_pk']).order_by('orden')

    def perform_create(self, serializer):
        formulario = Formulario.objects.get(pk=self.kwargs['formulario_pk'])
        serializer.save(formulario=formulario)


class OpcionRespuestaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar las Opciones de Respuesta de una Pregunta.
    Anidado bajo /formularios/{formulario_pk}/preguntas/{pregunta_pk}/opciones/
    """
    serializer_class = OpcionRespuestaSerializer
    permission_classes = [IsAdminOrDirectivo]

    def get_queryset(self):
        return OpcionRespuesta.objects.filter(pregunta_id=self.kwargs['pregunta_pk']).order_by('orden')

    def perform_create(self, serializer):
        pregunta = Pregunta.objects.get(pk=self.kwargs['pregunta_pk'])
        serializer.save(pregunta=pregunta)


class RespuestaUsuarioViewSet(viewsets.ViewSet):
    """
    ViewSet unificado para gestionar las respuestas a formularios.
    Toda la lógica ahora usa el modelo `RespuestaUsuario`, vinculado al `CustomUser`.
    - Usuarios autenticados pueden crear y ver sus propias respuestas.
    - Funcionarios/Admins pueden listar todas las respuestas y ver las de cualquier usuario.
    """
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action == 'list':
            self.permission_classes = [IsAdminOrFuncionario]
        elif self.action in ['retrieve', 'create']:
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()

    def list(self, request, formulario_pk=None):
        """
        Devuelve todas las respuestas para un formulario específico, agrupadas por usuario.
        Solo accesible para administradores y funcionarios.
        """
        respuestas = RespuestaUsuario.objects.filter(
            pregunta__formulario_id=formulario_pk
        ).select_related('usuario', 'pregunta').order_by('usuario__id')

        data = []
        # Agrupar respuestas por usuario
        for usuario, group in groupby(respuestas, key=attrgetter('usuario')):
            group_list = list(group)
            respuestas_serializadas = RespuestaUsuarioSerializer(group_list, many=True).data

            # Determinar el nombre a mostrar
            nombre_display = usuario.get_full_name() or usuario.username
            if hasattr(usuario, 'perfil_prestador') and usuario.perfil_prestador.nombre_negocio:
                nombre_display = usuario.perfil_prestador.nombre_negocio
            elif hasattr(usuario, 'perfil_artesano') and usuario.perfil_artesano.nombre_taller:
                nombre_display = usuario.perfil_artesano.nombre_taller

            data.append({
                'usuario_id': usuario.id,
                'nombre_display': nombre_display,
                'rol': usuario.get_role_display(),
                'fecha_ultima_respuesta': group_list[-1].fecha_respuesta if group_list else None,
                'respuestas': respuestas_serializadas
            })

        return Response(data)

    def retrieve(self, request, formulario_pk=None, usuario_pk=None):
        """
        Recupera las respuestas de un usuario específico para un formulario.
        - Un usuario solo puede ver sus propias respuestas (a menos que sea admin/funcionario).
        - Admins/Funcionarios pueden ver las respuestas de cualquier usuario.
        """
        user = request.user
        target_user_id = usuario_pk if usuario_pk is not None else user.id

        # Permitir a admin/funcionarios ver cualquier perfil
        if user.role not in [CustomUser.Role.ADMIN, CustomUser.Role.FUNCIONARIO_DIRECTIVO, CustomUser.Role.FUNCIONARIO_PROFESIONAL]:
            if str(target_user_id) != str(user.id):
                return Response({"error": "No tienes permiso para ver estas respuestas."}, status=status.HTTP_403_FORBIDDEN)

        try:
            # Si se especifica usuario_pk, se usa, si no, se usa el del usuario autenticado
            user_to_view = CustomUser.objects.get(pk=target_user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        respuestas = RespuestaUsuario.objects.filter(
            usuario=user_to_view,
            pregunta__formulario_id=formulario_pk
        ).order_by('pregunta__orden')

        if not respuestas.exists():
             # Devuelve una lista vacía si no hay respuestas, lo cual es un estado válido
            return Response([], status=status.HTTP_200_OK)

        serializer = RespuestaUsuarioSerializer(respuestas, many=True)
        return Response(serializer.data)

    def create(self, request, formulario_pk=None):
        """
        Crea o actualiza respuestas para un formulario para el usuario autenticado.
        La lógica se ha unificado; siempre se guardan como RespuestaUsuario.
        """
        serializer = RespuestaUsuarioCreateSerializer(
            data={'respuestas': request.data.get('respuestas'), 'formulario_id': formulario_pk},
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response({"status": "Respuestas guardadas con éxito."}, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class NotificacionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para que los usuarios vean sus notificaciones.
    """
    serializer_class = NotificacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Devuelve solo las notificaciones del usuario autenticado.
        """
        return self.request.user.notificaciones.all()

    @action(detail=False, methods=['post'], url_path='marcar-como-leidas')
    def marcar_como_leidas(self, request):
        """
        Marca todas las notificaciones no leídas del usuario como leídas.
        """
        request.user.notificaciones.filter(leido=False).update(leido=True)
        return Response(status=status.HTTP_204_NO_CONTENT)


class DetailedStatisticsView(views.APIView):
    """
    Vista para proporcionar estadísticas detalladas sobre la participación
    en formularios y el cumplimiento de normativas, consultadas por ID.
    """
    permission_classes = [IsAdminOrFuncionario]

    def get(self, request, *args, **kwargs):
        # --- Resumen General ---
        user_counts = CustomUser.objects.values('role').annotate(count=Count('id')).order_by('role')
        prestadores_total = PrestadorServicio.objects.count()
        artesanos_total = Artesano.objects.count()

        summary_data = {
            "usuarios": {"total": CustomUser.objects.count(), "por_rol": {item['role']: item['count'] for item in user_counts}},
            "prestadores": {"total": prestadores_total},
            "artesanos": {"total": artesanos_total},
            "publicaciones": {"total": Publicacion.objects.count()},
            "atractivos": {"total": AtractivoTuristico.objects.count()},
        }

        # --- Análisis de Participación en Formularios ---
        form_id = request.query_params.get('form_id')
        form_participation_data = {}
        if form_id:
            try:
                formulario = Formulario.objects.get(pk=form_id)
                # Usuarios que han respondido al menos una pregunta de este formulario
                usuarios_con_respuesta = RespuestaUsuario.objects.filter(pregunta__formulario=formulario).values_list('usuario_id', flat=True).distinct()

                # Determinar el universo de usuarios objetivo para este formulario (ej. todos los prestadores)
                # Esta lógica puede ser más compleja, pero por ahora asumimos que es para todos los prestadores/artesanos
                total_objetivo = prestadores_total + artesanos_total # Simplificación

                form_participation_data = {
                    'formulario_nombre': formulario.titulo,
                    'total_respuestas': len(usuarios_con_respuesta),
                    'total_objetivo': total_objetivo,
                    'tasa_participacion': (len(usuarios_con_respuesta) / total_objetivo * 100) if total_objetivo > 0 else 0
                }
            except Formulario.DoesNotExist:
                pass # No hacer nada si el form_id no es válido

        # --- Análisis de Cumplimiento por Ítem de Verificación ---
        item_id = request.query_params.get('item_id')
        compliance_analysis_data = {}
        if item_id:
            try:
                item = ItemVerificacion.objects.get(pk=item_id)
                respuestas_item = RespuestaItemVerificacion.objects.filter(item_original=item)

                cumplen = respuestas_item.filter(cumple=True).count()
                no_cumplen = respuestas_item.filter(cumple=False).count()

                # Evolución temporal del cumplimiento
                cumplimiento_por_ano = (
                    respuestas_item.annotate(ano=models.functions.ExtractYear('verificacion__fecha_visita'))
                    .values('ano')
                    .annotate(
                        total_cumplen=Count('pk', filter=models.Q(cumple=True)),
                        total_no_cumplen=Count('pk', filter=models.Q(cumple=False))
                    )
                    .order_by('ano')
                )

                compliance_analysis_data = {
                    'item_texto': item.texto_requisito,
                    'distribucion_actual': {'cumplen': cumplen, 'no_cumplen': no_cumplen},
                    'evolucion_temporal': list(cumplimiento_por_ano)
                }
            except ItemVerificacion.DoesNotExist:
                pass

        # --- Ensamblar Respuesta Final ---
        response_data = {
            'summary': summary_data,
            'form_participation': form_participation_data,
            'compliance_analysis': compliance_analysis_data,
        }

        return Response(response_data)


class PlantillaVerificacionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para listar y ver las plantillas de verificación.
    Solo los funcionarios y administradores pueden acceder a estas.
    """
    queryset = PlantillaVerificacion.objects.all().prefetch_related('items').order_by('nombre')
    permission_classes = [IsAdminOrFuncionario]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PlantillaVerificacionDetailSerializer
        return PlantillaVerificacionListSerializer


class VerificacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar las verificaciones de cumplimiento.
    - Funcionarios/Admins: Pueden iniciar, ver y guardar todas las verificaciones.
    - Prestadores: Pueden ver su propio historial de verificaciones.
    """
    queryset = Verificacion.objects.all().select_related(
        'plantilla_usada', 'prestador', 'funcionario_evaluador'
    ).prefetch_related('respuestas_items__item_original').order_by('-fecha_visita')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'iniciar_verificacion':
            return IniciarVerificacionSerializer
        if self.action == 'partial_update':
            return GuardarVerificacionSerializer
        if self.action == 'retrieve':
            return VerificacionDetailSerializer
        return VerificacionListSerializer

    def get_queryset(self):
        user = self.request.user
        if user.is_staff or user.role in [CustomUser.Role.ADMIN, CustomUser.Role.FUNCIONARIO_DIRECTIVO, CustomUser.Role.FUNCIONARIO_PROFESIONAL]:
            return super().get_queryset()
        if hasattr(user, 'perfil_prestador'):
            return super().get_queryset().filter(prestador=user.perfil_prestador)
        return Verificacion.objects.none()

    @action(detail=False, methods=['post'], url_path='iniciar')
    @transaction.atomic
    def iniciar_verificacion(self, request):
        """
        Crea una nueva instancia de Verificacion, pre-poblada con todos los ítems
        de la plantilla seleccionada, listos para ser llenados.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        plantilla = serializer.validated_data['plantilla_id']
        prestador = serializer.validated_data['prestador_id']
        funcionario = request.user

        # Crear la instancia de la verificación
        verificacion = Verificacion.objects.create(
            plantilla_usada=plantilla,
            prestador=prestador,
            funcionario_evaluador=funcionario,
            fecha_visita=timezone.now().date()
        )

        # Crear las respuestas por defecto para cada ítem de la plantilla
        items_de_plantilla = plantilla.items.all()
        respuestas_a_crear = [
            RespuestaItemVerificacion(
                verificacion=verificacion,
                item_original=item,
                cumple=False,
                justificacion=''
            ) for item in items_de_plantilla
        ]
        RespuestaItemVerificacion.objects.bulk_create(respuestas_a_crear)

        # Devolver la verificación recién creada y completamente poblada
        response_serializer = VerificacionDetailSerializer(verificacion)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        """
        Actualiza una verificación existente con las respuestas del formulario.
        La lógica de cálculo de puntajes está en el GuardarVerificacionSerializer.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated_instance = serializer.save()

        # Devolver la instancia actualizada con todos los detalles
        response_serializer = VerificacionDetailSerializer(updated_instance)
        return Response(response_serializer.data, status=status.HTTP_200_OK)

    def perform_create(self, serializer):
        # La creación se maneja a través de 'iniciar_verificacion'
        return Response({"error": "La creación directa no está permitida. Use la acción 'iniciar'."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def update(self, request, *args, **kwargs):
        return Response({"error": "La actualización completa (PUT) no está permitida. Use PATCH."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


def log_audit_action(user, instance, action, details=None):
    AuditLog.objects.create(
        user=user,
        content_object=instance,
        action=action,
        details=json.dumps(details) if details else None
    )

class PrestadorProfileView(generics.RetrieveUpdateAPIView):
    queryset = PrestadorServicio.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return PrestadorServicioUpdateSerializer
        return PrestadorServicioSerializer

    def get_object(self):
        try:
            return self.request.user.perfil_prestador
        except PrestadorServicio.DoesNotExist:
            from django.http import Http404
            raise Http404("El perfil de prestador no fue encontrado para este usuario.")


class ArtesanoProfileView(generics.RetrieveUpdateAPIView):
    queryset = Artesano.objects.all()
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return ArtesanoUpdateSerializer
        return ArtesanoSerializer

    def get_object(self):
        try:
            return self.request.user.artesano
        except Artesano.DoesNotExist:
            from django.http import Http404
            raise Http404("El perfil de artesano no fue encontrado para este usuario.")


class FeedbackProveedorView(generics.ListAPIView):
    serializer_class = FeedbackProveedorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        content_object = None
        if hasattr(user, 'perfil_prestador'):
            content_object = user.perfil_prestador
        elif hasattr(user, 'artesano'):
            content_object = user.artesano

        if content_object:
            content_type = ContentType.objects.get_for_model(content_object)
            return Sugerencia.objects.filter(
                content_type=content_type,
                object_id=content_object.pk,
                tipo_mensaje__in=[Sugerencia.TipoMensaje.QUEJA, Sugerencia.TipoMensaje.SUGERENCIA]
            ).order_by('-fecha_envio')
        return Sugerencia.objects.none()


class ImagenGaleriaView(generics.ListCreateAPIView):
    serializer_class = ImagenGaleriaSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return ImagenGaleria.objects.filter(prestador=self.request.user.perfil_prestador)

    def perform_create(self, serializer):
        serializer.save(prestador=self.request.user.perfil_prestador)


class ImagenGaleriaDetailView(generics.RetrieveDestroyAPIView):
    serializer_class = ImagenGaleriaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ImagenGaleria.objects.filter(prestador=self.request.user.perfil_prestador)


class ImagenArtesanoView(generics.ListCreateAPIView):
    serializer_class = ImagenArtesanoSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return ImagenArtesano.objects.filter(artesano=self.request.user.artesano)

    def perform_create(self, serializer):
        serializer.save(artesano=self.request.user.artesano)


class ImagenArtesanoDetailView(generics.RetrieveDestroyAPIView):
    serializer_class = ImagenArtesanoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ImagenArtesano.objects.filter(artesano=self.request.user.artesano)


class DocumentoLegalizacionView(generics.ListCreateAPIView):
    serializer_class = DocumentoLegalizacionSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return DocumentoLegalizacion.objects.filter(prestador=self.request.user.perfil_prestador)

    def perform_create(self, serializer):
        serializer.save(prestador=self.request.user.perfil_prestador)


class PrestadorRegisterView(RegisterView):
    serializer_class = PrestadorRegisterSerializer

class DocumentoLegalizacionDetailView(generics.RetrieveDestroyAPIView):
    serializer_class = DocumentoLegalizacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return DocumentoLegalizacion.objects.filter(prestador=self.request.user.perfil_prestador)


class PrestadorRegisterView(RegisterView):
    serializer_class = PrestadorRegisterSerializer

class TuristaRegisterView(RegisterView):
    serializer_class = TuristaRegisterSerializer


class ArtesanoRegisterView(RegisterView):
    serializer_class = ArtesanoRegisterSerializer


class ElementoGuardadoViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsTurista]

    def get_queryset(self):
        return ElementoGuardado.objects.filter(usuario=self.request.user)

    def get_serializer_class(self):
        if self.action == 'create':
            return ElementoGuardadoCreateSerializer
        return ElementoGuardadoSerializer

    def get_serializer_context(self):
        return {'request': self.request}


class ResenaViewSet(viewsets.ModelViewSet):
    queryset = Resena.objects.all().order_by('-fecha_creacion')

    def get_serializer_class(self):
        if self.action == 'create':
            return ResenaCreateSerializer
        return ResenaSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        elif self.action == 'create':
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAdminOrFuncionario]
        return super().get_permissions()

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list':
            queryset = queryset.filter(aprobada=True)
            content_type_str = self.request.query_params.get('content_type')
            object_id_str = self.request.query_params.get('object_id')
            if content_type_str and object_id_str:
                model_map = {'prestadorservicio': PrestadorServicio, 'artesano': Artesano}
                Model = model_map.get(content_type_str.lower())
                if Model:
                    try:
                        content_type = ContentType.objects.get_for_model(Model)
                        queryset = queryset.filter(content_type=content_type, object_id=object_id_str)
                    except ContentType.DoesNotExist:
                        return queryset.none()
        return queryset

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrFuncionario])
    def approve(self, request, pk=None):
        resena = self.get_object()
        resena.aprobada = True
        resena.save(update_fields=['aprobada'])
        return Response({'status': 'Reseña aprobada con éxito.'}, status=status.HTTP_200_OK)


class FelicitacionesPublicasView(generics.ListAPIView):
    queryset = Sugerencia.objects.filter(
        tipo_mensaje=Sugerencia.TipoMensaje.FELICITACION,
        es_publico=True
    ).order_by('-fecha_envio')
    serializer_class = FelicitacionPublicaSerializer
    permission_classes = [AllowAny]
    pagination_class = None


class SugerenciaViewSet(viewsets.mixins.CreateModelMixin, viewsets.GenericViewSet):
    queryset = Sugerencia.objects.all()
    serializer_class = SugerenciaSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        if self.request.user.is_authenticated:
            serializer.save(usuario=self.request.user)
        else:
            serializer.save()


class SugerenciaAdminViewSet(viewsets.ModelViewSet):
    queryset = Sugerencia.objects.all().order_by('-fecha_envio')
    serializer_class = SugerenciaAdminSerializer
    permission_classes = [IsAdminOrFuncionario]
    filter_backends = [OrderingFilter, SearchFilter]
    search_fields = ['mensaje', 'nombre_remitente', 'email_remitente', 'usuario__username']
    ordering_fields = ['fecha_envio', 'estado', 'tipo_mensaje']


class PublicacionListView(generics.ListAPIView):
    serializer_class = PublicacionListSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['titulo', 'contenido']
    ordering_fields = ['fecha_publicacion', 'titulo', 'fecha_evento_inicio']

    def get_queryset(self):
        config = SiteConfiguration.load()
        if not config.seccion_publicaciones_activa:
            return Publicacion.objects.none()
        queryset = Publicacion.objects.filter(estado=Publicacion.Status.PUBLICADO)
        if self.request.query_params.get('destacados'):
            limit = int(self.request.query_params.get('limit', 5))
            return queryset.filter(tipo='EVENTO', fecha_evento_inicio__gte=timezone.now()).order_by('fecha_evento_inicio')[:limit]
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date and end_date:
            queryset = queryset.filter(tipo='EVENTO', fecha_evento_inicio__lte=end_date, fecha_evento_fin__gte=start_date)
        tipos = self.request.query_params.get('tipo', None)
        if tipos:
            lista_tipos = [tipo.strip().upper() for tipo in tipos.split(',')]
            queryset = queryset.filter(tipo__in=lista_tipos)
        return queryset


class GaleriaListView(generics.ListAPIView):
    serializer_class = GaleriaItemSerializer
    permission_classes = [AllowAny]
    pagination_class = None

    def get_queryset(self):
        imagenes = ImagenAtractivo.objects.filter(atractivo__es_publicado=True).select_related('atractivo')
        videos = Video.objects.all()
        unified_list = []
        for img in imagenes:
            unified_list.append({
                'id': f'imagen_{img.id}', 'tipo': 'imagen', 'url': self.request.build_absolute_uri(img.imagen.url),
                'thumbnail_url': self.request.build_absolute_uri(img.imagen.url), 'titulo': img.atractivo.nombre,
                'descripcion': img.alt_text or f"Imagen de {img.atractivo.nombre}"
            })
        for vid in videos:
            video_id = vid.url_youtube.split('v=')[-1].split('&')[0]
            thumbnail = f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg"
            unified_list.append({
                'id': f'video_{vid.id}', 'tipo': 'video', 'url': vid.url_youtube,
                'thumbnail_url': thumbnail, 'titulo': vid.titulo, 'descripcion': vid.descripcion
            })
        import random
        random.shuffle(unified_list)
        return unified_list


class PublicacionDetailView(generics.RetrieveAPIView):
    serializer_class = PublicacionDetailSerializer
    permission_classes = [AllowAny]
    queryset = Publicacion.objects.filter(estado=Publicacion.Status.PUBLICADO)
    lookup_field = 'slug'


class ConsejoConsultivoListView(generics.ListAPIView):
    queryset = ConsejoConsultivo.objects.filter(es_publicado=True).order_by('-fecha_publicacion')
    serializer_class = ConsejoConsultivoSerializer
    permission_classes = [AllowAny]


class VideoListView(generics.ListAPIView):
    queryset = Video.objects.all().order_by('-fecha_publicacion')
    serializer_class = VideoSerializer
    permission_classes = [AllowAny]


class CategoriaPrestadorListView(generics.ListAPIView):
    queryset = CategoriaPrestador.objects.all().order_by('nombre')
    serializer_class = CategoriaPrestadorSerializer
    permission_classes = [AllowAny]
    pagination_class = None


class RubroArtesanoListView(generics.ListAPIView):
    queryset = RubroArtesano.objects.all().order_by('nombre')
    serializer_class = RubroArtesanoSerializer
    permission_classes = [AllowAny]
    pagination_class = None


class PrestadorServicioPublicListView(generics.ListAPIView):
    serializer_class = PrestadorServicioPublicListSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['nombre_negocio', 'descripcion']
    ordering_fields = ['-puntuacion_total', 'nombre_negocio']

    def get_queryset(self):
        config = SiteConfiguration.load()
        if not config.seccion_prestadores_activa:
            return PrestadorServicio.objects.none()
        queryset = PrestadorServicio.objects.filter(aprobado=True).order_by('-puntuacion_total', 'nombre_negocio')
        categoria_slug = self.request.query_params.get('categoria', None)
        if categoria_slug:
            queryset = queryset.filter(categoria__slug=categoria_slug)
        return queryset


class PrestadorServicioPublicDetailView(generics.RetrieveAPIView):
    queryset = PrestadorServicio.objects.filter(aprobado=True)
    serializer_class = PrestadorServicioPublicDetailSerializer
    permission_classes = [AllowAny]
    lookup_field = 'pk'


class ArtesanoPublicListView(generics.ListAPIView):
    serializer_class = ArtesanoPublicListSerializer
    permission_classes = [AllowAny]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['nombre_taller', 'nombre_artesano', 'descripcion']
    ordering_fields = ['-puntuacion_total', 'nombre_taller']

    def get_queryset(self):
        queryset = Artesano.objects.filter(aprobado=True).order_by('-puntuacion_total', 'nombre_taller')
        rubro_slug = self.request.query_params.get('rubro', None)
        if rubro_slug:
            queryset = queryset.filter(rubro__slug=rubro_slug)
        return queryset


class ArtesanoPublicDetailView(generics.RetrieveAPIView):
    queryset = Artesano.objects.filter(aprobado=True)
    serializer_class = ArtesanoPublicDetailSerializer
    permission_classes = [AllowAny]
    lookup_field = 'pk'


class AtractivoTuristicoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión completa de Atractivos Turísticos.
    Permite el acceso público para lectura y acceso restringido para escritura
    según las reglas de CanManageAtractivos.
    """
    queryset = AtractivoTuristico.objects.all().order_by('nombre')
    permission_classes = [CanManageAtractivos]
    parser_classes = [MultiPartParser, FormParser]
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return AtractivoTuristicoWriteSerializer
        elif self.action == 'retrieve':
            return AtractivoTuristicoDetailSerializer
        return AtractivoTuristicoListSerializer

    def get_queryset(self):
        user = self.request.user
        # Si el usuario no está autenticado o no es staff, solo ve los publicados.
        if not user.is_authenticated or not user.is_staff:
            config = SiteConfiguration.load()
            if not config.seccion_atractivos_activa:
                return AtractivoTuristico.objects.none()
            return self.queryset.filter(es_publicado=True)

        # El staff puede ver todos los atractivos.
        return self.queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

    def perform_create(self, serializer):
        # El autor se asigna en el serializador
        serializer.save()
    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrDirectivo])
    def approve(self, request, slug=None):
        """
        Acción para que un Admin o Directivo apruebe un atractivo turístico.
        """
        atractivo = self.get_object()
        atractivo.es_publicado = True
        atractivo.save(update_fields=['es_publicado'])
        return Response({'status': 'Atractivo turístico aprobado y publicado con éxito.'}, status=status.HTTP_200_OK)


class RutaTuristicaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para la visualización de Rutas Turísticas.
    Permite el acceso público para lectura.
    """
    queryset = RutaTuristica.objects.filter(es_publicado=True).prefetch_related(
        'atractivos', 'prestadores'
    ).order_by('nombre')
    permission_classes = [AllowAny]
    lookup_field = 'slug'

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return RutaTuristicaDetailSerializer
        return RutaTuristicaListSerializer


class LocationListView(views.APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        locations = []

        # Obtener Prestadores con latitud y longitud válidas
        prestadores = PrestadorServicio.objects.filter(
            aprobado=True, latitud__isnull=False, longitud__isnull=False
        ).select_related('categoria')
        for p in prestadores:
            locations.append({
                'id': f'prestador_{p.id}',
                'nombre': p.nombre_negocio,
                'lat': p.latitud,
                'lng': p.longitud,
                'tipo': p.categoria.slug if p.categoria else 'prestador',
                'url_detalle': f'/directorio/prestador/{p.id}'
            })

        # Obtener Atractivos con latitud y longitud válidas
        atractivos = AtractivoTuristico.objects.filter(
            es_publicado=True, latitud__isnull=False, longitud__isnull=False
        )
        for a in atractivos:
            locations.append({
                'id': f'atractivo_{a.id}',
                'nombre': a.nombre,
                'lat': a.latitud,
                'lng': a.longitud,
                'tipo': f'atractivo_{a.categoria_color.lower()}',
                'url_detalle': f'/atractivos/{a.slug}'
            })

        # Obtener Artesanos con latitud y longitud válidas
        artesanos = Artesano.objects.filter(
            aprobado=True, latitud__isnull=False, longitud__isnull=False
        )
        for art in artesanos:
            locations.append({
                'id': f'artesano_{art.id}',
                'nombre': art.nombre_taller,
                'lat': art.latitud,
                'lng': art.longitud,
                'tipo': 'artesano',
                'url_detalle': f'/directorio/artesano/{art.id}'
            })

        serializer = LocationSerializer(locations, many=True)
        return Response(serializer.data)


class AdminArtesanoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión completa de Artesanos por parte de los administradores.
    """
    queryset = Artesano.objects.select_related('usuario', 'rubro').prefetch_related('galeria_imagenes').order_by('-fecha_creacion')
    permission_classes = [IsAdminOrFuncionario]
    filter_backends = [OrderingFilter, SearchFilter]
    search_fields = ['nombre_taller', 'nombre_artesano', 'usuario__email']
    ordering_fields = ['fecha_creacion', 'nombre_taller', 'aprobado']

    def get_serializer_class(self):
        if self.action == 'list':
            return AdminArtesanoListSerializer
        return AdminArtesanoDetailSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        aprobado_param = self.request.query_params.get('aprobado')
        if aprobado_param is not None:
            aprobado = aprobado_param.lower() == 'true'
            queryset = queryset.filter(aprobado=aprobado)
        return queryset

    @action(detail=True, methods=['patch'])
    def approve(self, request, pk=None):
        artesano = self.get_object()
        artesano.aprobado = True
        artesano.save(update_fields=['aprobado'])
        return Response({'status': 'Artesano aprobado con éxito.'}, status=status.HTTP_200_OK)


class AdminUsuarioListView(generics.ListAPIView):
    """
    Vista para que los administradores y funcionarios obtengan una lista de todos
    los usuarios relevantes (excluyendo turistas).
    """
    serializer_class = UsuarioListSerializer
    permission_classes = [IsAdminOrFuncionario]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['username', 'first_name', 'last_name', 'email']
    ordering_fields = ['username', 'role']

    def get_queryset(self):
        # Excluimos a los turistas ya que no llenan formularios de caracterización.
        return CustomUser.objects.exclude(role=CustomUser.Role.TURISTA).order_by('username')


class AdminPrestadorViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión completa de Prestadores de Servicio por parte de los administradores.
    """
    queryset = PrestadorServicio.objects.select_related('usuario', 'categoria').prefetch_related('galeria_imagenes', 'documentos_legalizacion').order_by('-fecha_creacion')
    permission_classes = [IsAdminOrFuncionario]
    filter_backends = [OrderingFilter, SearchFilter]
    search_fields = ['nombre_negocio', 'usuario__email']
    ordering_fields = ['fecha_creacion', 'nombre_negocio', 'aprobado']

    def get_serializer_class(self):
        if self.action == 'list':
            return AdminPrestadorListSerializer
        return AdminPrestadorDetailSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        aprobado_param = self.request.query_params.get('aprobado')
        if aprobado_param is not None:
            aprobado = aprobado_param.lower() == 'true'
            queryset = queryset.filter(aprobado=aprobado)
        return queryset

    @action(detail=True, methods=['patch'])
    def approve(self, request, pk=None):
        prestador = self.get_object()
        prestador.aprobado = True
        prestador.save(update_fields=['aprobado'])
        return Response({'status': 'Prestador aprobado con éxito.'}, status=status.HTTP_200_OK)


class AdminPublicacionViewSet(viewsets.ModelViewSet):
    queryset = Publicacion.objects.all().order_by('-fecha_publicacion')
    serializer_class = AdminPublicacionSerializer
    permission_classes = [IsAdminOrFuncionario]
    parser_classes = [MultiPartParser, FormParser]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            instance = self.get_object()
            if instance.tipo == Publicacion.Tipo.CAPACITACION:
                return CapacitacionDetailSerializer
        if self.action == 'registrar_asistencia':
            return RegistrarAsistenciaSerializer
        return super().get_serializer_class()

    def perform_create(self, serializer):
        user = self.request.user
        # Default state is BORRADOR. Let actions handle state changes.
        serializer.save(autor=user)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrFuncionario], url_path='submit-for-approval')
    def submit_for_approval(self, request, pk=None):
        publicacion = self.get_object()
        user = request.user

        if publicacion.estado != Publicacion.Status.BORRADOR:
            return Response({'error': 'La publicación no está en estado de borrador.'}, status=status.HTTP_400_BAD_REQUEST)

        if user.role == CustomUser.Role.FUNCIONARIO_PROFESIONAL:
            publicacion.estado = Publicacion.Status.PENDIENTE_DIRECTIVO
            publicacion.save()
            return Response({'status': 'Publicación enviada para aprobación del Directivo.'})

        elif user.role == CustomUser.Role.FUNCIONARIO_DIRECTIVO:
            publicacion.estado = Publicacion.Status.PENDIENTE_ADMIN
            publicacion.save()
            return Response({'status': 'Publicación enviada para aprobación del Administrador.'})

        return Response({'error': 'No tienes permiso para realizar esta acción.'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrDirectivo], url_path='approve')
    def approve(self, request, pk=None):
        publicacion = self.get_object()
        user = request.user

        if user.role == CustomUser.Role.FUNCIONARIO_DIRECTIVO:
            if publicacion.estado == Publicacion.Status.PENDIENTE_DIRECTIVO:
                publicacion.estado = Publicacion.Status.PENDIENTE_ADMIN
                publicacion.save()
                return Response({'status': 'Publicación aprobada y enviada al Administrador.'})
            else:
                return Response({'error': 'Esta publicación no está pendiente de su aprobación.'}, status=status.HTTP_400_BAD_REQUEST)

        if user.role == CustomUser.Role.ADMIN:
            if publicacion.estado in [Publicacion.Status.PENDIENTE_ADMIN, Publicacion.Status.PENDIENTE_DIRECTIVO]:
                publicacion.estado = Publicacion.Status.PUBLICADO
                publicacion.save()
                return Response({'status': 'Publicación aprobada y publicada.'})
            else:
                return Response({'error': 'Esta publicación no está pendiente de aprobación.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'error': 'No tienes permiso para aprobar.'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrDirectivo], url_path='reject')
    def reject(self, request, pk=None):
        publicacion = self.get_object()
        if publicacion.estado not in [Publicacion.Status.PENDIENTE_DIRECTIVO, Publicacion.Status.PENDIENTE_ADMIN]:
            return Response({'error': 'La publicación no está pendiente de aprobación.'}, status=status.HTTP_400_BAD_REQUEST)

        publicacion.estado = Publicacion.Status.BORRADOR
        publicacion.save()
        # Optionally, add a reason for rejection in a new field or log. For now, just change state.
        return Response({'status': 'La publicación ha sido devuelta a borrador.'})

    @action(detail=True, methods=['post'], url_path='registrar-asistencia', permission_classes=[IsAdminOrFuncionario])
    def registrar_asistencia(self, request, pk=None):
        """
        Registra la asistencia de una lista de usuarios a esta capacitación.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = serializer.save()
        return Response(result, status=status.HTTP_200_OK)


class ContenidoMunicipioViewSet(viewsets.ModelViewSet):
    queryset = ContenidoMunicipio.objects.all().order_by('orden')
    serializer_class = ContenidoMunicipioSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdminOrFuncionario]
        return super().get_permissions()

    def perform_create(self, serializer):
        instance = serializer.save(actualizado_por=self.request.user)
        log_audit_action(self.request.user, instance, AuditLog.Action.CONTENIDO_CREATE, details={'titulo': instance.titulo})

    def perform_update(self, serializer):
        instance = serializer.save(actualizado_por=self.request.user)
        log_audit_action(self.request.user, instance, AuditLog.Action.CONTENIDO_UPDATE, details={'titulo': instance.titulo})

    def perform_destroy(self, instance):
        details = {'titulo': instance.titulo}
        log_audit_action(self.request.user, instance, AuditLog.Action.CONTENIDO_DELETE, details)
        instance.delete()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({'request': self.request})
        return context


class HomePageComponentViewSet(viewsets.ModelViewSet):
    serializer_class = HomePageComponentSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        if self.action == 'list' and not self.request.user.is_authenticated:
            return HomePageComponent.objects.filter(is_active=True).order_by('order')
        return HomePageComponent.objects.all().order_by('order')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdminOrFuncionario]
        return super().get_permissions()

    def perform_create(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.COMPONENT_CREATE, details={'title': instance.title})

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.COMPONENT_UPDATE, details={'title': instance.title})

    def perform_destroy(self, instance):
        details = {'title': instance.title}
        log_audit_action(self.request.user, instance, AuditLog.Action.COMPONENT_DELETE, details)
        instance.delete()

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrFuncionario])
    def reorder(self, request, *args, **kwargs):
        ordered_ids = request.data.get('ordered_ids', [])
        if not isinstance(ordered_ids, list):
            return Response({"error": "Se esperaba una lista de IDs."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            with transaction.atomic():
                for index, component_id in enumerate(ordered_ids):
                    HomePageComponent.objects.filter(pk=component_id).update(order=index)
            first_component = HomePageComponent.objects.filter(pk__in=ordered_ids).first()
            if first_component:
                log_audit_action(request.user, first_component, AuditLog.Action.COMPONENT_REORDER, details={"info": f"Se reordenaron {len(ordered_ids)} componentes."})
            return Response({"status": "Componentes reordenados con éxito."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Error al reordenar: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PaginaInstitucionalViewSet(viewsets.ModelViewSet):
    queryset = PaginaInstitucional.objects.all()
    serializer_class = PaginaInstitucionalSerializer
    parser_classes = [MultiPartParser, FormParser]
    lookup_field = 'slug'

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdminOrFuncionario]
        return super().get_permissions()

    def perform_create(self, serializer):
        instance = serializer.save(actualizado_por=self.request.user)
        log_audit_action(self.request.user, instance, AuditLog.Action.CONTENIDO_CREATE, details={'nombre_pagina': instance.nombre})

    def perform_update(self, serializer):
        instance = serializer.save(actualizado_por=self.request.user)
        log_audit_action(self.request.user, instance, AuditLog.Action.CONTENIDO_UPDATE, details={'nombre_pagina': instance.nombre})


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all().select_related('user').prefetch_related('content_object')
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdmin]
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_class = AuditLogFilter
    ordering_fields = ['timestamp']
    search_fields = ['user__username', 'details', 'action']


class UserViewSet(viewsets.ModelViewSet):
    serializer_class = AdminUserSerializer
    permission_classes = [IsAdminOrFuncionarioForUserManagement]

    def get_queryset(self):
        user = self.request.user
        if user.role == CustomUser.Role.ADMIN:
            return CustomUser.objects.all().order_by('username')
        if user.role in [CustomUser.Role.FUNCIONARIO_DIRECTIVO, CustomUser.Role.FUNCIONARIO_PROFESIONAL]:
            return CustomUser.objects.filter(
                models.Q(role=CustomUser.Role.PRESTADOR) |
                models.Q(role=CustomUser.Role.ARTESANO) |
                models.Q(role=CustomUser.Role.TURISTA) |
                models.Q(pk=user.pk)
            ).order_by('username').distinct()
        return CustomUser.objects.none()

    def perform_create(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.USER_CREATE, details={'username': instance.username, 'role': instance.role})

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.USER_UPDATE, details={'username': instance.username, 'role': instance.role})

    def perform_destroy(self, instance):
        details = {'username': instance.username, 'role': instance.role}
        log_audit_action(self.request.user, instance, AuditLog.Action.USER_DELETE, details)
        instance.delete()


class SiteConfigurationView(generics.RetrieveUpdateAPIView):
    serializer_class = SiteConfigurationSerializer

    def get_object(self):
        return SiteConfiguration.load()

    def get_permissions(self):
        if self.request.method in ['GET', 'HEAD', 'OPTIONS']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.SITE_CONFIG_UPDATE, details={'updated_fields': list(serializer.validated_data.keys())})


class MenuItemViewSet(viewsets.ModelViewSet):
    serializer_class = MenuItemSerializer

    def get_queryset(self):
        if self.action == 'list':
            return MenuItem.objects.filter(parent__isnull=True).order_by('orden')
        return MenuItem.objects.all().order_by('orden')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdmin]
        return super().get_permissions()

    @action(detail=False, methods=['post'], url_path='reorder')
    @transaction.atomic
    def reorder(self, request, *args, **kwargs):
        structured_menu = request.data
        def update_items_recursive(items, parent=None):
            for index, item_data in enumerate(items):
                item_id = item_data.get('id')
                if not item_id: continue
                try:
                    menu_item = MenuItem.objects.get(id=item_id)
                    menu_item.orden = index
                    menu_item.parent = parent
                    menu_item.save(update_fields=['orden', 'parent'])
                    if 'children' in item_data and item_data['children']:
                        update_items_recursive(item_data['children'], parent=menu_item)
                except MenuItem.DoesNotExist:
                    continue
        try:
            update_items_recursive(structured_menu)
            log_audit_action(request.user, None, AuditLog.Action.MENU_UPDATE, details={"info": "Estructura del menú reordenada."})
            return Response({"status": "Menú reordenado con éxito."}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"Error al reordenar el menú: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.MENU_CREATE, details={'name': instance.nombre, 'url': instance.url})

    def perform_update(self, serializer):
        instance = serializer.save()
        log_audit_action(self.request.user, instance, AuditLog.Action.MENU_UPDATE, details={'name': instance.nombre, 'url': instance.url})

    def perform_destroy(self, instance):
        details = {'name': instance.nombre, 'url': instance.url}
        log_audit_action(self.request.user, instance, AuditLog.Action.MENU_DELETE, details)
        instance.delete()


class HechoHistoricoViewSet(viewsets.ModelViewSet):
    serializer_class = HechoHistoricoSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        if self.action in ['list', 'retrieve']:
            return HechoHistorico.objects.filter(es_publicado=True).order_by('ano')
        return HechoHistorico.objects.all().order_by('ano')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [AllowAny]
        else:
            self.permission_classes = [IsAdminOrFuncionario]
        return super().get_permissions()


def run_agent_in_background(task_id):
    from agents.corps.turismo_coronel import get_turismo_coronel_graph
    from django.conf import settings

    try:
        task = AgentTask.objects.select_related('user').get(id=task_id)

        user_context = None
        general_order = task.command
        is_guest = task.user is None

        if is_guest:
            # Crear un contexto de "invitado"
            guest_user = CustomUser(
                username="Invitado",
                role=CustomUser.Role.TURISTA,
                ai_provider=CustomUser.AIProvider.OPENAI, # Usar el proveedor por defecto del sistema
                api_key=settings.OPENAI_API_KEY_GLOBAL # Usar la clave global segura
            )
            user_context = guest_user
            general_order = f"Soy un visitante nuevo en la plataforma. Mi pregunta es: '{task.command}'. Ayúdame y, si es relevante, invítame a registrarme para una experiencia completa."
        else:
            # Usar el contexto del usuario autenticado
            user = task.user
            if not user.api_key or not user.ai_provider:
                raise ValueError("El usuario no ha configurado su proveedor de IA o su clave de API personal.")
            user_context = user

        task.status = AgentTask.Status.RUNNING
        task.save()

        coronel_agent = get_turismo_coronel_graph()
        config = {"configurable": {"thread_id": f"task_{task_id}"}}

        result = asyncio.run(coronel_agent.ainvoke({
            "general_order": general_order,
            "app_context": {"user": user_context, "is_guest": is_guest}
        }, config=config))

        task.report = result.get("final_report", "La misión concluyó sin un informe detallado.")
        task.status = AgentTask.Status.COMPLETED
        task.save()

    except Exception as e:
        if 'task' in locals():
            task.report = f"Error crítico durante la ejecución de la misión: {str(e)}"
            task.status = AgentTask.Status.FAILED
            task.save()
        print(f"Error en el hilo del agente para la tarea {task_id}: {e}")


class AgentCommandView(views.APIView):
    permission_classes = [AllowAny] # Permitir acceso a invitados

    def post(self, request, *args, **kwargs):
        serializer = AgentCommandSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        command = serializer.validated_data['orden']
        user = request.user if request.user.is_authenticated else None

        task = AgentTask.objects.create(user=user, command=command, status=AgentTask.Status.PENDING)

        thread = threading.Thread(target=run_agent_in_background, args=(task.id,))
        thread.start()

        return Response({"message": "Comando recibido. El agente ha sido desplegado.", "task_id": task.id}, status=status.HTTP_202_ACCEPTED)


class AgentTaskStatusView(generics.RetrieveAPIView):
    queryset = AgentTask.objects.all()
    serializer_class = AgentTaskSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'

    def get_queryset(self):
        return AgentTask.objects.filter(user=self.request.user)


class LLMKeysView(generics.RetrieveUpdateAPIView):
    serializer_class = LLMKeysSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user


class AnalyticsDataView(views.APIView):
    """
    Vista para generar los datos del panel de análisis principal.
    Calcula un conjunto de métricas clave para la gestión y estrategia turística.
    """
    permission_classes = [IsAdminOrFuncionario]

    def get(self, request, *args, **kwargs):
        # --- 1. Resumen General (Tarjetas) ---
        total_prestadores = PrestadorServicio.objects.filter(aprobado=True).count()
        total_artesanos = Artesano.objects.filter(aprobado=True).count()
        total_publicaciones = Publicacion.objects.filter(estado=Publicacion.Status.PUBLICADO).count()
        avg_rating = Resena.objects.filter(aprobada=True).aggregate(avg=models.Avg('calificacion'))['avg'] or 0

        summary_cards = {
            "total_prestadores": { "value": total_prestadores, "label": "Prestadores Activos" },
            "total_artesanos": { "value": total_artesanos, "label": "Artesanos Activos" },
            "total_publicaciones": { "value": total_publicaciones, "label": "Publicaciones" },
            "calificacion_promedio": { "value": f"{avg_rating:.1f}", "label": "Calificación Promedio" },
        }

        # --- 2. Distribución de Usuarios por Rol (Gráfico de Torta) ---
        user_roles_data = list(CustomUser.objects.values('role').annotate(value=Count('id')).order_by('-value'))
        # Mapear los nombres de rol para que sean más legibles
        role_display_map = dict(CustomUser.Role.choices)
        for item in user_roles_data:
            item['name'] = role_display_map.get(item['role'], item['role'])

        # --- 3. Proveedores por Categoría (Gráfico de Barras) ---
        prestadores_por_categoria = list(
            PrestadorServicio.objects.filter(aprobado=True, categoria__isnull=False)
            .values('categoria__nombre')
            .annotate(value=Count('id'))
            .order_by('-value')
            .rename(name='categoria__nombre')
        )
        artesanos_por_rubro = list(
            Artesano.objects.filter(aprobado=True, rubro__isnull=False)
            .values('rubro__nombre')
            .annotate(value=Count('id'))
            .order_by('-value')
            .rename(name='rubro__nombre')
        )

        # --- 4. Evolución de Registros (Gráfico de Líneas) ---
        registros_por_mes = (
            CustomUser.objects
            .annotate(month=TruncDay('date_joined'))
            .values('month')
            .annotate(
                prestadores=Count('id', filter=models.Q(role=CustomUser.Role.PRESTADOR)),
                artesanos=Count('id', filter=models.Q(role=CustomUser.Role.ARTESANO)),
                turistas=Count('id', filter=models.Q(role=CustomUser.Role.TURISTA))
            )
            .order_by('month')
        )
        # Formatear para recharts
        evolucion_registros = [
            {
                "date": item['month'].strftime('%Y-%m-%d'),
                "Prestadores": item['prestadores'],
                "Artesanos": item['artesanos'],
                "Turistas": item['turistas'],
            }
            for item in registros_por_mes if item['month'] > timezone.now() - timedelta(days=365) # Último año
        ]

        # --- 5. Top 5 Proveedores por Puntuación (Tabla) ---
        top_prestadores = list(
            PrestadorServicio.objects.filter(aprobado=True)
            .order_by('-puntuacion_total')
            .values('nombre_negocio', 'puntuacion_total')[:5]
        )
        top_artesanos = list(
            Artesano.objects.filter(aprobado=True)
            .order_by('-puntuacion_total')
            .values('nombre_taller', 'puntuacion_total')[:5]
        )

        # --- 6. Métricas de Gestión Adicionales (Como las que ya existían) ---
        num_capacitaciones = Publicacion.objects.filter(tipo=Publicacion.Tipo.CAPACITACION, estado=Publicacion.Status.PUBLICADO).count()
        num_prestadores_formalizados = DocumentoLegalizacion.objects.filter(nombre_documento__icontains='RNT').values_list('prestador_id', flat=True).distinct().count()

        additional_metrics = {
            "capacitaciones_realizadas": { "value": num_capacitaciones, "label": "Capacitaciones Realizadas" },
            "prestadores_con_rnt": { "value": num_prestadores_formalizados, "label": "Prestadores con RNT" },
        }

        # --- Ensamblar Respuesta Final ---
        data = {
            'summary_cards': summary_cards,
            'user_roles_distribution': user_roles_data,
            'providers_by_category': {
                'prestadores': prestadores_por_categoria,
                'artesanos': artesanos_por_rubro,
            },
            'registration_over_time': evolucion_registros,
            'top_providers': {
                'prestadores': top_prestadores,
                'artesanos': top_artesanos,
            },
            'additional_metrics': additional_metrics,
        }

        return Response(data, status=status.HTTP_200_OK)


class ScoringRuleViewSet(viewsets.ViewSet):
    """
    ViewSet para gestionar la única instancia de ScoringRule (Singleton).
    Solo los administradores pueden ver y actualizar las reglas de puntuación.
    """
    permission_classes = [IsAdmin]

    def list(self, request):
        """
        Obtiene la instancia única de ScoringRule. La crea si no existe.
        """
        rules, created = ScoringRule.objects.get_or_create(pk=1)
        serializer = ScoringRuleSerializer(rules)
        return Response(serializer.data)

    def update(self, request, pk=None):
        """
        Actualiza la instancia única de ScoringRule.
        """
        try:
            rules = ScoringRule.objects.get(pk=1)
        except ScoringRule.DoesNotExist:
            return Response({"error": "No se encontraron las reglas de puntuación."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ScoringRuleSerializer(rules, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AIConfigView(generics.RetrieveUpdateAPIView):
    """
    Gestiona la configuración de IA personal para el usuario autenticado.
    Permite ver el proveedor de IA actual y actualizar tanto el proveedor
    como la clave de API personal (la clave es de solo escritura).
    """
    serializer_class = AIConfigSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        """
        Devuelve el objeto de usuario para el cual se está gestionando la configuración,
        que es siempre el usuario que realiza la solicitud.
        """
        return self.request.user

    def get(self, request, *args, **kwargs):
        """
        Al obtener la configuración, la clave de API no se muestra.
        El serializador se encarga de esto.
        """
        return self.retrieve(request, *args, **kwargs)

    def put(self, request, *args, **kwargs):
        """
        Permite la actualización completa de la configuración.
        """
        return self.update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        """
        Permite la actualización parcial de la configuración.
        """
        return self.partial_update(request, *args, **kwargs)


from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

class ExportExcelView(views.APIView):
    """
    Vista para exportar datos de los modelos principales a un archivo Excel.
    """
    permission_classes = [IsAdminOrFuncionario]

    def get(self, request, *args, **kwargs):
        # Crear un libro de trabajo de Excel en memoria
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Eliminar la hoja por defecto

        # --- Hoja de Prestadores de Servicios ---
        prestadores_qs = PrestadorServicio.objects.all().select_related('usuario', 'categoria')
        prestadores_sheet = workbook.create_sheet(title="Prestadores de Servicios")

        # Encabezados
        prestador_headers = [
            'ID', 'Nombre del Negocio', 'Categoría', 'Descripción', 'Teléfono', 'Email',
            'Facebook', 'Instagram', 'TikTok', 'WhatsApp', 'Dirección', 'Latitud', 'Longitud',
            'Aprobado', 'Puntuación Total'
        ]
        prestadores_sheet.append(prestador_headers)

        # Datos
        for prestador in prestadores_qs:
            prestadores_sheet.append([
                str(prestador.id),
                str(prestador.nombre_negocio or ''),
                str(prestador.categoria.nombre if prestador.categoria else ''),
                str(prestador.descripcion or ''),
                str(prestador.telefono or ''),
                str(prestador.email_contacto or ''),
                str(prestador.red_social_facebook or ''),
                str(prestador.red_social_instagram or ''),
                str(prestador.red_social_tiktok or ''),
                str(prestador.red_social_whatsapp or ''),
                str(prestador.direccion or ''),
                str(prestador.latitud or ''),
                str(prestador.longitud or ''),
                "Sí" if prestador.aprobado else "No",
                str(prestador.puntuacion_total)
            ])

        # --- Hoja de Artesanos ---
        artesanos_qs = Artesano.objects.all().select_related('usuario', 'rubro')
        artesanos_sheet = workbook.create_sheet(title="Artesanos")

        # Encabezados
        artesano_headers = [
            'ID', 'Nombre del Taller', 'Nombre del Artesano', 'Rubro', 'Descripción', 'Teléfono',
            'Email', 'Facebook', 'Instagram', 'TikTok', 'WhatsApp', 'Dirección', 'Latitud',
            'Longitud', 'Aprobado', 'Puntuación Total'
        ]
        artesanos_sheet.append(artesano_headers)

        # Datos
        for artesano in artesanos_qs:
            artesanos_sheet.append([
                str(artesano.id),
                str(artesano.nombre_taller or ''),
                str(artesano.nombre_artesano or ''),
                str(artesano.rubro.nombre if artesano.rubro else ''),
                str(artesano.descripcion or ''),
                str(artesano.telefono or ''),
                str(artesano.email_contacto or ''),
                str(artesano.red_social_facebook or ''),
                str(artesano.red_social_instagram or ''),
                str(artesano.red_social_tiktok or ''),
                str(artesano.red_social_whatsapp or ''),
                str(artesano.direccion or ''),
                str(artesano.latitud or ''),
                str(artesano.longitud or ''),
                "Sí" if artesano.aprobado else "No",
                str(artesano.puntuacion_total)
            ])

        # --- Hoja de Atractivos Turísticos ---
        atractivos_qs = AtractivoTuristico.objects.all().select_related('autor')
        atractivos_sheet = workbook.create_sheet(title="Atractivos Turísticos")

        # Encabezados
        atractivo_headers = [
            'ID', 'Nombre', 'Descripción', 'Cómo Llegar', 'Dirección', 'Latitud', 'Longitud',
            'Categoría', 'Publicado', 'Autor'
        ]
        atractivos_sheet.append(atractivo_headers)

        # Datos
        for atractivo in atractivos_qs:
            atractivos_sheet.append([
                str(atractivo.id),
                str(atractivo.nombre or ''),
                str(atractivo.descripcion or ''),
                str(atractivo.como_llegar or ''),
                str(atractivo.direccion or ''),
                str(atractivo.latitud or ''),
                str(atractivo.longitud or ''),
                str(atractivo.get_categoria_color_display() or ''),
                "Sí" if atractivo.es_publicado else "No",
                str(atractivo.autor.username if atractivo.autor else '')
            ])

        # Crear la respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_turismo.xlsx"'
        workbook.save(response)

        return response